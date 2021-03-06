# -*- coding:UTF-8 -*-
import os.path
import pyodbc
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, colors, Font
from colorama import init, Fore, Style

init()

#Read configuration info
if not os.path.exists('AppSettings.xml'):
    print(Fore.RED + "Cannot find configuration file.")
    exit()

databaseConnectionPool = {}

def getDatabaseConfiguration(databaseName):
    result = {'driver':'','server':'', 'database':databaseName, 'userName':'', 'password':'', 'trustedConnection':''}
    
    configurationContent = '\n'.join(open('AppSettings.xml').readlines())
    configurationDocument = BeautifulSoup(configurationContent, features='xml')
    databaseSettingsNode = configurationDocument.find('databaseSettings')
    if databaseSettingsNode is not None:
        databaseConfigurationNode = databaseSettingsNode.find('database', databaseName=databaseName)
        if databaseConfigurationNode is not None:
            if databaseConfigurationNode.get('driver') is not None:
                result['driver'] = databaseConfigurationNode.get('driver')
            if databaseConfigurationNode.get('server') is not None:
                result['server'] = databaseConfigurationNode.get('server')
            if databaseConfigurationNode.get('userName') is not None:
                result['userName'] = databaseConfigurationNode.get('userName')
            if databaseConfigurationNode.get('password') is not None:
                result['password'] = databaseConfigurationNode.get('password')
            if databaseConfigurationNode.get('trustedConnection') is not None:
                result['trustedConnection'] = databaseConfigurationNode.get('trustedConnection')

    return result

def getDatabaseConnection(configuration):
    if configuration is not None:
        if configuration['trustedConnection'] == 'true':
            return pyodbc.connect('Driver={driver};'
                'Server={server};'
                'Database={database};Trusted_Connection=yes'.format(driver = configuration['driver'], server = configuration['server'],
                                                                    database = configuration['database']))
        else:
            return pyodbc.connect('Driver={driver};'
                'Server={server};'
                'Database={database};'
                'UID={username};'
                'PWD={password}'.format(driver = configuration['driver'], server = configuration['server'], database = configuration['database'],
                                        username = configuration['userName'], password = configuration['password']))
        

eudDatabaseConfiguration = getDatabaseConfiguration('EUD')
eudConn = getDatabaseConnection(eudDatabaseConfiguration)

def getTruckCenterConnection(truckCenterID):
    sql = '''select 
                tc.SqlDb,
                s.IPAddress
            from adtbTruckCenters as tc
                inner join adtbServers as s on tc.ServerName = s.ServerName
            where TruckCenterID = {tcID}'''.format(tcID = truckCenterID)
    dbCursor = eudConn.cursor()
    dbCursor.execute(sql)
    dbRow = dbCursor.fetchone()

    truckCenterDatabaseConfiguration = getDatabaseConfiguration('TruckCenter')
    
    if dbRow != None:
        if truckCenterID in databaseConnectionPool.keys():
            return databaseConnectionPool[truckCenterID]
        else:
            truckCenterDatabaseConfiguration['server'] = dbRow[1]
            truckCenterDatabaseConfiguration['database'] = dbRow[0]
            newConnection = getDatabaseConnection(truckCenterDatabaseConfiguration)
            databaseConnectionPool[truckCenterID] = newConnection
            return newConnection
    else:
        return None

checkSql = '''select * from
(select 
s.FileID,
s.ProcessDate,
s.ProcessResult,
s.MessageID,
s.FileLength,
s.Data.value('declare namespace n0="http://finance.group.volvo.com/vendorinvoicedetails/1_0";
(/n0:SyncVendorInvoiceDetails/VendorInvoiceDetails/VendorInvoiceHeader/CompanyCode)[1]', 'nvarchar(20)') AS CompanyCode,
s.Data
from satbSourceData s
where s.FileType = 'VendorInvoiceDetails') as t
where FileID in ('150523', '150525', '150526')
order by T.ProcessDate desc'''

sapDatabaseConfiguration = getDatabaseConfiguration('SAP')
conn = getDatabaseConnection(sapDatabaseConfiguration)
print(Fore.GREEN + 'Connected to SAP database.\n')
print(Style.RESET_ALL)
cursor = conn.cursor()
cursor.execute(checkSql)

print('Data fetched from SAP database.\n')

wb = openpyxl.workbook.Workbook()
overallSheet = wb.active
overallSheet.title = 'Overall'
overallSheet['A1'] = 'FileID'
overallSheet['B1'] = 'ProcessDate'
overallSheet['C1'] = 'ProcessResult'
overallSheet['D1'] = 'MessageID'
overallSheet['E1'] = 'FileLength'
overallSheet['F1'] = 'CompanyCode'

#Read configuration
configurationContent = '\n'.join(open('AppSettings.xml').readlines())
configurationDocument = BeautifulSoup(configurationContent, features='xml')

headerRowDisplayStyle = {'fontColor':'FFFFFF', 'backgroundColor':'000080'}
highLightRowDisplayStyle = {'fontColor':'FFFFFF', 'backgroundColor':'FFFF00'}
uiConfiguration = configurationDocument.find('uiSettings')
if not uiConfiguration is None:
    headerRowConfiguration = uiConfiguration.find('headerRow')
    if not headerRowConfiguration is None:
        if not headerRowConfiguration.get('fontColor') is None:
            headerRowDisplayStyle['fontColor'] = headerRowConfiguration.get('fontColor')
        if not headerRowConfiguration.get('backgroundColor') is None:
            headerRowDisplayStyle['backgroundColor'] = headerRowConfiguration.get('backgroundColor')
    highLightRowConfiguration = uiConfiguration.find('highLightRow')
    if not highLightRowConfiguration is None:
        if not highLightRowConfiguration.get('fontColor') is None:
            highLightRowDisplayStyle['fontColor'] = highLightRowConfiguration.get('fontColor')
        if not highLightRowConfiguration.get('backgroundColor') is None:
            highLightRowDisplayStyle['backgroundColor'] = highLightRowConfiguration.get('backgroundColor')

validTypeOfInvoiceValues = []
validDocumentTypeValues = []
validationConfiguration = configurationDocument.find('validationSettings')
if not validationConfiguration is None:
    typeOfInvoiceConfiguration = validationConfiguration.find('typeOfInvoice')
    if not typeOfInvoiceConfiguration is None:
        for item in typeOfInvoiceConfiguration.find_all('item'):
            if not item.get('value') is None:
                validTypeOfInvoiceValues.append(item.get('value'))
    
    documentTypeConfiguration = validationConfiguration.find('documentType')
    if not documentTypeConfiguration is None:
        for item in documentTypeConfiguration.find_all('item'):
            if not item.get('value') is None:
                validDocumentTypeValues.append(item.get('value'))

headerRowFill = PatternFill(start_color=headerRowDisplayStyle['backgroundColor'],
                   end_color=headerRowDisplayStyle['backgroundColor'],
                   fill_type='solid')
highLightRowFill = PatternFill(start_color=highLightRowDisplayStyle['backgroundColor'],
                   end_color=highLightRowDisplayStyle['backgroundColor'],
                   fill_type='solid')
headerRowFont = Font(color=headerRowDisplayStyle['fontColor'])
highLightRowFont = Font(color=highLightRowDisplayStyle['fontColor'])

maxColumn = overallSheet.max_column
for i in range(1, maxColumn + 1):
    overallSheet.cell(row = 1, column = i).fill = headerRowFill
    overallSheet.cell(row = 1, column = i).font = headerRowFont

rowIndex = 2
sourceFileRows = cursor.fetchall()
for row in sourceFileRows:
    fileID = row[0]
    processDate = row[1]
    processResult = row[2]
    messageID = row[3]
    fileLength = row[4]
    companyCode = row[5]
    message = row[6]

    overallSheet['A' + str(rowIndex)] = fileID
    overallSheet['B' + str(rowIndex)] = processDate
    overallSheet['C' + str(rowIndex)] = processResult
    overallSheet['D' + str(rowIndex)] = messageID
    overallSheet['E' + str(rowIndex)] = fileLength
    overallSheet['F' + str(rowIndex)] = companyCode

    print('Analyzing content of File #{fID}\n'.format(fID = fileID))

    rowIndex += 1
    fileSheet = wb.create_sheet(title=str(fileID), index=1)
    fileSheet['A1'] = 'Accounting Document' # DocumentNumber
    fileSheet['B1'] = 'Year' # FiscalYear
    fileSheet['C1'] = 'Doc. Date' # InvoiceDate
    fileSheet['D1'] = 'Pstng Date' # PostingDate
    fileSheet['E1'] = 'Doc. No.' # InvoiceDocumentNumber
    fileSheet['F1'] = 'PO Number' # PurchasingDocumentNumber
    fileSheet['G1'] = 'P++' # OrderReference
    fileSheet['H1'] = 'PO line item' # PurchasingDocumentNumberItem
    fileSheet['I1'] = 'Item PO Number' # Purchase Order No in Invoice Item 
    fileSheet['J1'] = 'Vendor #' # VendorNumber
    fileSheet['K1'] = 'PO Currency' # CurrencyCode
    fileSheet['L1'] = 'Type of Invoice' # TypeOfInvoice
    fileSheet['M1'] = 'Document Type' # DocumentType
    fileSheet['N1'] = 'Quantity' # Quantity
    fileSheet['O1'] = 'GrossValue' # GrossValue
    fileSheet['P1'] = 'Status'
    fileSheet['Q1'] = 'Authorised Value'
    fileSheet['R1'] = 'Invoice Value'
    fileSheet['S1'] = 'Supplier'
    fileSheet['T1'] = 'Supplier Length'
    fileSheet['U1'] = 'PO Index Sync'

    maxColumn = fileSheet.max_column
    for i in range(1, maxColumn + 1):
        fileSheet.cell(row = 1, column = i).fill = headerRowFill
        fileSheet.cell(row = 1, column = i).font = headerRowFont

    fileRowIndex = 2

    soup = BeautifulSoup(message, features='xml')
    for header in soup.find_all('VendorInvoiceHeader'):
        documentNumber = header.find('DocumentNumber').get_text()
        fiscalYear = header.find('FiscalYear').get_text()
        invoiceDate = header.find('InvoiceDate').get_text()
        postingDate = header.find('PostingDate').get_text()
        invoiceDocumentNumber = header.find('InvoiceDocumentNumber').get_text()
        orderReference = header.find('OrderReference').get_text()
        vendorNumber = header.find('VendorNumber').get_text()
        currencyCode = header.find('CurrencyCode').get_text()
        typeOfInvoice = header.find('TypeOfInvoice').get_text()
        documentType = header.find('DocumentType').get_text()
        invoiceNumber = header.find('InvoiceNumber').get_text()

        print('\tFound DocumentNumber {docNumber} with TypeOfInvoice as {typeOfInv}, DocumentType as {docType}\n'.format(docNumber = documentNumber, typeOfInv = typeOfInvoice, docType = documentType))

        for item in header.parent.find_all('VendorInvoiceItem'):
            purchasingDocumentNumber = item.find('PurchasingDocumentNumber').get_text()
            purchasingDocumentNumberItem = item.find('PurchasingDocumentNumberItem').get_text()
            itemOrderReferenceNo = item.find('OrderReference').get_text()
            quantity = item.find('Quantity').get_text()
            grossValue = item.find('GrossValue').get_text()
            
            fileSheet['A' + str(fileRowIndex)] = documentNumber
            fileSheet['B' + str(fileRowIndex)] = fiscalYear
            fileSheet['C' + str(fileRowIndex)] = invoiceDate
            fileSheet['D' + str(fileRowIndex)] = postingDate
            fileSheet['E' + str(fileRowIndex)] = invoiceDocumentNumber
            fileSheet['F' + str(fileRowIndex)] = purchasingDocumentNumber
            fileSheet['G' + str(fileRowIndex)] = orderReference
            fileSheet['H' + str(fileRowIndex)] = purchasingDocumentNumberItem
            fileSheet['I' + str(fileRowIndex)] = itemOrderReferenceNo
            fileSheet['J' + str(fileRowIndex)] = vendorNumber
            fileSheet['K' + str(fileRowIndex)] = currencyCode
            fileSheet['L' + str(fileRowIndex)] = typeOfInvoice
            if typeOfInvoice not in validTypeOfInvoiceValues:
                fileSheet['L' + str(fileRowIndex)].fill = highLightRowFill
                fileSheet['L' + str(fileRowIndex)].font = highLightRowFont
            fileSheet['M' + str(fileRowIndex)] = documentType
            if documentType not in validDocumentTypeValues:
                fileSheet['M' + str(fileRowIndex)].fill = highLightRowFill
                fileSheet['M' + str(fileRowIndex)].font = highLightRowFont
            fileSheet['N' + str(fileRowIndex)] = quantity
            fileSheet['O' + str(fileRowIndex)] = grossValue

            #Check Invoice from Truck Center Database
            truckCenterID = ''
            checkInvoiceSql = """select 
                                    item.TruckCenterID
                                from satbVendorInvoices as inv
                                    left join satbVendorInvoiceItems as item on inv.InvoiceID = item.InvoiceID
                                where inv.OrderReference = '{orderRef}'
                                    and inv.InvoiceNumber = '{invNo}'
                                    and item.PurchasingDocumentNumberItem = '{purDocNoItem}'""".format(orderRef = orderReference, invNo = invoiceNumber, purDocNoItem = purchasingDocumentNumberItem)
            invoiceCursor = conn.cursor()
            invoiceCursor.execute(checkInvoiceSql)
            invoiceRow = invoiceCursor.fetchone()
            if invoiceRow != None:
                truckCenterID = invoiceRow[0]
                if truckCenterID is None:
                    print(Fore.RED + '\tMissing Index for {orderRef}'.format(orderRef = orderReference))
                    print(Style.RESET_ALL)
                    continue

                checkTruckCenterSql = """select 
                                                po.AuthorisedValue,
                                                po.InvoiceValue,
                                                status.Status,
                                                po.GDS_Supplier
                                        from ppvwAllOrderSummary as po
                                            left join pptbStatus as status on po.StatusID = status.StatusID
                                        where po.ActualOrderNo = '{poNumber}'""".format(poNumber = itemOrderReferenceNo)
                truckCenterConn = getTruckCenterConnection(truckCenterID)
                truckCenterCursor = truckCenterConn.cursor()
                truckCenterCursor.execute(checkTruckCenterSql)
                truckCenterRow = truckCenterCursor.fetchone()
                if truckCenterRow != None:
                    fileSheet['P' + str(fileRowIndex)] = truckCenterRow[2]
                    fileSheet['Q' + str(fileRowIndex)] = truckCenterRow[0]
                    fileSheet['R' + str(fileRowIndex)] = truckCenterRow[1]
                    fileSheet['S' + str(fileRowIndex)] = truckCenterRow[3]
                    fileSheet['T' + str(fileRowIndex)] = len(truckCenterRow[3])

                    checkPOIndexSql = """select 
                                                case when count(*) > 0 then 'YES' else 'NO' end as POIndexCreated
                                        from satbPurchaseOrderIndex
                                        where PurchaseOrderNo = '{orderRef}'
                                                and CompanyCode = '{companyCode}'
                                                and TruckCenterID = {tcID}""".format(orderRef = itemOrderReferenceNo, companyCode = companyCode, tcID = truckCenterID)
                    
                    checkIndexConn = getDatabaseConnection(sapDatabaseConfiguration)
                    poIndexCursor = checkIndexConn.cursor()
                    poIndexCursor.execute(checkPOIndexSql)
                    poIndexRow = poIndexCursor.fetchone()
                    if poIndexRow != None:
                        fileSheet['U' + str(fileRowIndex)] = poIndexRow[0]
                    checkIndexConn.close()

            cornerCell = fileSheet['A2']
            fileSheet.freeze_panes = cornerCell
            fileRowIndex += 1

wb.save('xmlMessage.xlsx')
print(Fore.GREEN + 'finished.\n')
conn.close()
print('Disconnected from SAP database.\n')
eudConn.close()
print('Disconnected from EUD database.\n')
for connectionKey in databaseConnectionPool.keys():
    databaseConnectionPool[connectionKey].close()
    print('Disconnected from Truck Center {truckCenterID} database.\n'.format(truckCenterID = connectionKey))
print(Style.RESET_ALL)

